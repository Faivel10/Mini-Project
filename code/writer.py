import os
from openpyxl import load_workbook

def check_and_update_excel(file_path, sheet_name, workbook):
    try:
        sheet = workbook[sheet_name]
        with open(file_path, 'r') as file:
            print("opened file: " + file_path)
            lines = file.readlines()
        

        #set all weights to 1 in output
        for i, row_cells in enumerate(sheet.iter_rows()):
            if i == 0:
                sheet['C' + str(i + 1)] = 'W'
                continue
            sheet['C' + str(i + 1)] = 1

        for line in lines:
            x, y = line.strip().split(' ')
            found = False
            for i, row_cells in enumerate(sheet.iter_rows()):
                if i == 0:
                    continue
                if float(row_cells[0].value) == float(x):
                    sheet['C' + str(i + 1)] = 2
                    #print("saved in index: ", i + 1)
                    found = True
                    break
            if not found:
                print("failed! didnt find point with x: ", x)
        workbook.save("faivel_dragatsky.xlsx")
        print(f"Processed " + sheet_name + file_path)

    except Exception as e:
        print(f"Error processing {file_path}: {e}")

def process_files_in_folder(folder_path, excel_file_path):
    workbook = load_workbook(excel_file_path)
    for filename in os.listdir(folder_path):
            sheet_name = os.path.splitext(filename)[0]
            print(sheet_name)
            if sheet_name.split('-')[1] in workbook.sheetnames:
                file_path = os.path.join(folder_path, filename)
                check_and_update_excel(file_path, sheet_name.split('-')[1], workbook)

folder_path = "./parsed-output"
excel_file_path = "./data.xlsx"
process_files_in_folder(folder_path, excel_file_path)
